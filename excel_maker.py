import openpyxl


class Excel_Maker:
    def __init__(self, noextractor, option_dict):
        super().__init__()

        self.noextractor = noextractor
        self.option_dict = option_dict

    def start(self):
        workbook = openpyxl.Workbook()
        sheet_name_list = ['제목', '본문', '제목+본문']

        for sheet_name in sheet_name_list:
            sheet = workbook.create_sheet(sheet_name)
            if sheet_name == '제목':
                self.make_sheet(sheet, self.noextractor.title_counter)
            elif sheet_name == '본문':
                self.make_sheet(sheet, self.noextractor.body_counter)
            elif sheet_name == '제목+본문':
                dict = {}
                for key, value in self.noextractor.title_counter.items():
                    if key in self.noextractor.body_counter.keys():
                        dict[key] = value + self.noextractor.body_counter[key]
                        del self.noextractor.body_counter[key]
                    else:
                        dict[key] = value

                for key, value in self.noextractor.body_counter.items():
                    dict[key] = value

                self.make_sheet(sheet, dict)

            row = 0
            for cnt in sheet.rows:
                row += 1
                if cnt[1].value == "총 빈도 수":
                    continue
                noun_sum = 0
                for i in range(2, len(cnt)):
                    if cnt[i].value is not None:
                        noun_sum += cnt[i].value
                sheet.cell(row=row, column=2, value=noun_sum)

        workbook.remove(workbook['Sheet'])
        filename = "[{0}_{1}_{2}]results.xlsx".format(self.option_dict['pages'],
                                                      self.option_dict['start'], self.option_dict['end'])
        workbook.save(filename=filename)

        print("\"{0}\" 파일로 저장되었습니다.".format(filename))

    def make_sheet(self, sheet, counter):
        row = 2
        column = 3
        sheet.cell(row=1, column=2, value="총 빈도 수")
        word_list = {}
        for date in counter:
            sheet.cell(row=1, column=column, value=date)
            for word, count in sorted(counter[date].items(),
                                      key=lambda t: t[1], reverse=True):
                if word in word_list.keys():
                    value = sheet.cell(row=word_list[word], column=column).value
                    if value is not None:
                        value += count
                    else:
                        value = count
                    sheet.cell(row=word_list[word], column=column, value=value)
                else:
                    sheet.cell(row=row, column=1, value=word)
                    sheet.cell(row=row, column=column, value=count)
                    word_list[word] = row
                    row += 1
            column += 1
